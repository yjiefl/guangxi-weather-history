"""
数据导出服务
负责将天气数据导出为Excel和CSV格式
遵循单一职责原则和接口隔离原则
"""
import logging
import pandas as pd
from io import BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class DataExporter:
    """
    数据导出器类
    负责将数据导出为不同格式
    """
    
    def __init__(self):
        """初始化数据导出器"""
        logger.info("数据导出器初始化完成")
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        fields: List[str] = None,
        include_summary: bool = True,
        city_name: str = None
    ) -> bytes:
        """
        导出数据为Excel格式
        
        Args:
            data: 数据列表
            filename: 文件名
            fields: 要导出的字段列表，None表示导出所有字段
            include_summary: 是否包含汇总表
            
        Returns:
            Excel文件的字节流
        """
        try:
            # 转换为DataFrame
            df = self._format_data(data, fields, city_name)
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认的工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 添加数据表
            ws_data = wb.create_sheet('天气数据')
            
            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)
            
            # 设置表头样式
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws_data[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 自动调整列宽
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # 添加汇总表
            if include_summary and len(df) > 0:
                self._add_summary_sheet(wb, df)
            
            # 保存到字节流
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"导出Excel成功: {filename}, 共 {len(df)} 条记录")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        fields: List[str] = None,
        city_name: str = None
    ) -> bytes:
        """
        导出数据为CSV格式
        
        Args:
            data: 数据列表
            filename: 文件名
            fields: 要导出的字段列表，None表示导出所有字段
            
        Returns:
            CSV文件的字节流
        """
        try:
            # 转换为DataFrame
            df = self._format_data(data, fields, city_name)
            
            # 转换为CSV
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')  # utf-8-sig支持Excel打开中文
            output.seek(0)
            
            logger.info(f"导出CSV成功: {filename}, 共 {len(df)} 条记录")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            raise
    
    def _format_data(
        self,
        data: List[Dict[str, Any]],
        fields: List[str] = None,
        city_name: str = None
    ) -> pd.DataFrame:
        """
        格式化数据为DataFrame
        
        Args:
            data: 数据列表
            fields: 要包含的字段列表
            
        Returns:
            格式化后的DataFrame
        """
        if not data:
            return pd.DataFrame()
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 添加城市信息
        if city_name:
            df['city'] = city_name
        
        # 拆分日期和时间
        if 'datetime' in df.columns:
            df['datetime_dt'] = pd.to_datetime(df['datetime'])
            df['date'] = df['datetime_dt'].dt.date
            df['time'] = df['datetime_dt'].dt.strftime('%H:%M')
        
        # 天气代码转换 (Item 17)
        if 'weather_code' in df.columns:
            weather_map = {
                0: '晴朗', 1: '晴到多云', 2: '多云', 3: '阴天', 45: '雾', 
                48: '沉积雾', 51: '小毛毛雨', 53: '毛毛雨', 55: '大毛毛雨',
                61: '小雨', 63: '中雨', 65: '大雨', 71: '小雪', 73: '中雪', 
                75: '大雪', 80: '阵雨', 81: '中阵雨', 82: '大阵雨', 95: '雷阵雨'
            }
            df['weather_code'] = df['weather_code'].map(lambda x: f"{int(x)} ({weather_map.get(int(x), '未知')})" if pd.notnull(x) else x)
        
        # 如果指定了字段，确保包含我们新增的辅助字段
        # 按照固定顺序排列核心字段，增加导出的整齐度 (Item 2 改进)
        helper_cols = ['city', 'date', 'time']
        core_order = [
            'temperature_2m', 'relative_humidity_2m', 'dew_point_2m',
            'precipitation', 'rain', 'snowfall', 'surface_pressure', 'cloud_cover',
            'wind_speed_10m', 'wind_direction_10m', 'wind_gusts_10m',
            'wind_speed_100m', 'wind_direction_100m',
            'shortwave_radiation', 'direct_radiation', 'diffuse_radiation', 'direct_normal_irradiance',
            'evapotranspiration', 'soil_temperature_0_to_7cm', 'soil_moisture_0_to_7cm',
            'weather_code'
        ]
        
        final_cols = [c for c in helper_cols if c in df.columns]
        for c in core_order:
            if c in df.columns:
                if fields is None or c in fields:
                    final_cols.append(c)
        
        # 严格限制字段：只导出 core_order 中定义的或明确请求的字段 (Item 2 & 13)
        # 不再通过循环 df.columns 来添加数据库中多余的空闲字段（如已弃用的 80m 风速等）
        
        df = df[final_cols]
        
        # 重命名列为中文
        column_mapping = self._get_column_mapping()
        df = df.rename(columns=column_mapping)
        
        # 删除临时的辅助列
        if 'datetime_dt' in df.columns:
            df = df.drop(columns=['datetime_dt'])
        
        return df
    def _get_column_mapping(self) -> Dict[str, str]:
        """
        获取列名映射（英文到中文）
        
        Returns:
            列名映射字典
        """
        return {
            'city': '城市',
            'date': '日期',
            'time': '时间',
            'datetime': '日期时间',
            'temperature_2m': '温度(°C)',
            'relative_humidity_2m': '相对湿度(%)',
            'dew_point_2m': '露点温度(°C)',
            'precipitation': '降水量(mm)',
            'rain': '降雨量(mm)',
            'snowfall': '降雪量(cm)',
            'surface_pressure': '地面气压(hPa)',
            'cloud_cover': '云量(%)',
            'wind_speed_10m': '10米风速(km/h)',
            'wind_direction_10m': '10米风向(°)',
            'wind_gusts_10m': '10米阵风(km/h)',
            'wind_speed_100m': '100米风速(km/h)',
            'wind_direction_100m': '100米风向(°)',
            'shortwave_radiation': '短波辐射(W/m²)',
            'direct_radiation': '直接辐射(W/m²)',
            'diffuse_radiation': '散射辐射(W/m²)',
            'direct_normal_irradiance': '直接法向辐照度(W/m²)',
            'evapotranspiration': '蒸发蒸腾量(mm)',
            'soil_temperature_0_to_7cm': '土壤温度(°C)',
            'soil_moisture_0_to_7cm': '土壤湿度(m³/m³)',
            'weather_code': '天气代码',
        }
    
    def _add_summary_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """
        添加汇总表到Excel工作簿
        
        Args:
            workbook: Excel工作簿对象
            df: 数据DataFrame
        """
        try:
            ws_summary = workbook.create_sheet('数据汇总')
            
            # 计算统计数据
            summary_data = []
            
            # 数值列
            numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
            
            for col in numeric_columns:
                if col != '日期时间':
                    stats = {
                        '数据项': col,
                        '平均值': round(df[col].mean(), 2) if not df[col].isna().all() else None,
                        '最大值': round(df[col].max(), 2) if not df[col].isna().all() else None,
                        '最小值': round(df[col].min(), 2) if not df[col].isna().all() else None,
                        '总和': round(df[col].sum(), 2) if not df[col].isna().all() else None,
                    }
                    summary_data.append(stats)
            
            # 创建汇总DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # 写入汇总数据
            for r in dataframe_to_rows(summary_df, index=False, header=True):
                ws_summary.append(r)
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 自动调整列宽
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_summary.column_dimensions[column_letter].width = adjusted_width

            # 新增：按日汇总表 (Item 15)
            if '日期' in df.columns:
                self._add_daily_summary_sheet(workbook, df)
            
            logger.debug("添加汇总表成功")
            
        except Exception as e:
            logger.error(f"添加汇总表失败: {e}")
            
    def _add_daily_summary_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """添加每日汇总表"""
        try:
            ws_daily = workbook.create_sheet('每日汇总')
            
            # 数值列（排除日期和城市）
            numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
            
            # 基础分组字段
            group_cols = ['日期']
            if '城市' in df.columns:
                group_cols.append('城市')
            
            # 按日汇总（均值）
            daily_df = df.groupby(group_cols)[numeric_cols].mean().reset_index()
            
            # 特殊处理：降水量、辐射和蒸发量应该是总和 (Item 2 修复)
            sum_cols = [
                '降水量(mm)', '降雨量(mm)', '蒸发蒸腾量(mm)',
                '短波辐射(W/m²)', '直接辐射(W/m²)', '散射辐射(W/m²)', 
                '直接法向辐照度(W/m²)'
            ]
            for col in sum_cols:
                if col in daily_df.columns:
                    # 确保是按日期和城市分组重新计算总和
                    daily_sum = df.groupby(group_cols)[col].sum().reset_index()
                    # 按照索引对齐，将均值替换为总和
                    daily_df[col] = daily_sum[col]
            
            # 写入数据
            for r in dataframe_to_rows(daily_df, index=False, header=True):
                ws_daily.append(r)
                
            # 设置样式（同上）
            header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            for cell in ws_daily[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # 自动列宽
            for column in ws_daily.columns:
                ws_daily.column_dimensions[column[0].column_letter].width = 15

        except Exception as e:
            logger.error(f"添加每日汇总表失败: {e}")
